from openpyxl import Workbook, load_workbook

def bssid_does_exist(bssid) -> bool:
    # Load the workbook
    wb = load_workbook(filename="./wifi-05-28-2022.xlsx")

    # Use this sheet!!
    ws = wb['simple']

    # Iterate over every cell in column A
    for cell in ws['A']:
        # Check if value matches
        if cell.value == bssid:
            return True

print(bssid_does_exist("fewf"))
